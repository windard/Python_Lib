# coding=utf-8

import openpyxl


filename = 'sku.xlsx'


def readExel():
    inwb = openpyxl.load_workbook(filename)  # 读文件
    sheetnames = inwb.get_sheet_names()  # 获取读文件中所有的sheet，通过名字的方式
    ws = inwb.get_sheet_by_name(sheetnames[0])  # 获取第一个sheet内容

    # 获取sheet的最大行数和列数
    rows = ws.max_row
    cols = ws.max_column
    for r in range(1, rows):
        for c in range(1, cols):
            print(str(ws.cell(row=r, column=c).value))
        if r == 10:
            break


def writeExcel():
    outwb = openpyxl.Workbook()  # 打开一个将写的文件
    outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
    for row in range(1, 70000):
        for col in range(1, 4):

            # outws.cell(row=row, column=col).value = row*2  # 写文件
            outws.cell(row=row, column=col, value=row * 2)  # 写文件
        print(row)
    outwb.save(filename)  # 一定要记得保存


if __name__ == '__main__':
    # writeExcel()
    readExel()
